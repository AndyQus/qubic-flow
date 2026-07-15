"""Excel workbook generation for the Bestandsverlauf feature.

Generates one .xlsx per series (hourly/daily/weekly), each a full clone of
the user's ledger template structure:

  Ledger           overview: one row per capture, newly added QUBIC per
                   wallet column, price per 1 billion QUBIC in EUR/USD,
                   per-owner EUR value columns with formulas
  Ledger<Owner>    one sheet per owner: event-level ledger
  Transfer         internal transfers between own wallets (- source / + dest)
  MyLedgerCSV      flat transaction list

Everything is rebuilt from the local database on every export — the files
are pure output artifacts, no personal data is baked into the code.
"""
import io
import logging
import os
import re
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from ..config import settings
from ..database import SessionLocal
from ..models.balance_snapshot import BalanceSnapshot, SnapshotAnnotation
from ..models.event import Event
from ..models.wallet import Wallet

logger = logging.getLogger(__name__)

BLN = 1_000_000_000

FMT_AMOUNT = '#,##0;[RED]\\-#,##0'
FMT_EUR = '#,##0.00\\ [$€-407];[RED]\\-#,##0.00\\ [$€-407]'
FMT_USD = '[$$-409]#,##0.00;[RED]\\-[$$-409]#,##0.00'
FMT_DATE = 'dd/mm/yy'
FMT_RATE = '0.0000000000'

FILE_NAMES = {
    "de": {"hourly": "stuendlich", "daily": "taeglich", "weekly": "woechentlich"},
    "en": {"hourly": "hourly", "daily": "daily", "weekly": "weekly"},
}

I18N = {
    "de": {
        "title": {"hourly": "Bestandsverlauf Stündlich", "daily": "Bestandsverlauf Täglich",
                  "weekly": "Bestandsverlauf Wöchentlich"},
        "date": "Datum", "epoch": "Epoch", "from": "Von", "to": "Bis",
        "eur_per_bln": " €  = 1Bln", "usd_per_bln": "$= 1Bln", "why": "why",
        "info": "Informationen", "notes": "Anmerkungen", "note": "Notiz",
        "total": "Gesamt", "prev_year": "Vorjahr", "cur_year": "Aktuelles Jahr",
        "no_owner": "OhneBesitzer", "transfers": "Transfers",
        "balance_sheet": "Bestand", "balance_title": "Bestand je Wallet (laut Qubic-Netzwerk zum Erfassungszeitpunkt)",
        "tick": "Tick", "value_eur": "Wert €", "value_usd": "Wert $",
        "residual_note": "Events (Balance-Differenz abzügl. TX)",
        "cat_internal": "verschoben", "cat_reward": "Reward", "cat_in": "Eingang",
        "cat_out": "Ausgang", "cat_manual": "manuell",
        "csv_headers": ["Datum", "Token", "Anzahl", "Preis in anderer Währung", "Preis in Dollar",
                        "Typ", "Asset", "Plattform", "Betrag", "Buchung", "Besitzer", "Text",
                        "publicId", "tx"],
        "typ_sc": "Smart Contract", "typ_tx": "Transfer",
        "text_sc": "Passives Einkommen über SmartContract",
    },
    "en": {
        "title": {"hourly": "Balance History Hourly", "daily": "Balance History Daily",
                  "weekly": "Balance History Weekly"},
        "date": "Date", "epoch": "Epoch", "from": "From", "to": "To",
        "eur_per_bln": " €  = 1Bln", "usd_per_bln": "$= 1Bln", "why": "why",
        "info": "Information", "notes": "Notes", "note": "Note",
        "total": "Total", "prev_year": "Previous year", "cur_year": "Current year",
        "no_owner": "NoOwner", "transfers": "Transfers",
        "balance_sheet": "Bestand", "balance_title": "Balance per wallet (per Qubic network at capture time)",
        "tick": "Tick", "value_eur": "Value €", "value_usd": "Value $",
        "residual_note": "Events (balance delta minus TX)",
        "cat_internal": "internal", "cat_reward": "Reward", "cat_in": "Incoming",
        "cat_out": "Outgoing", "cat_manual": "manual",
        "csv_headers": ["Date", "Token", "Amount", "Price EUR", "Price USD",
                        "Type", "Asset", "Platform", "Value", "Booking", "Owner", "Text",
                        "publicId", "tx"],
        "typ_sc": "Smart Contract", "typ_tx": "Transfer",
        "text_sc": "Passive income via smart contract",
    },
}


def _parse_iso(value: str | None) -> datetime | None:
    if not value:
        return None
    try:
        return datetime.fromisoformat(value.replace("Z", "+00:00")).replace(tzinfo=None)
    except Exception:
        return None


def _to_local(dt: datetime | None) -> datetime | None:
    """Naive UTC -> naive local time for Excel date cells. Excel has no
    timezone concept, so dates are written in the server's local timezone
    (TZ environment variable); the UI converts in the browser instead."""
    if dt is None:
        return None
    return dt.replace(tzinfo=timezone.utc).astimezone().replace(tzinfo=None)


def _series_baseline(db: Session, kind: str) -> datetime | None:
    """First capture (auto/manual) of the series. Rows before it lie outside
    the recorded period — same rule as build_owner_rows."""
    first = (
        db.query(BalanceSnapshot.captured_at)
        .filter(
            BalanceSnapshot.kind == kind,
            BalanceSnapshot.trigger.in_(("auto", "manual")),
        )
        .order_by(BalanceSnapshot.captured_at, BalanceSnapshot.id)
        .limit(1)
        .scalar()
    )
    return _parse_iso(first)


def _sheet_name(base: str, taken: set) -> str:
    name = re.sub(r"[\[\]:*?/\\']", "", base).strip()[:31] or "Sheet"
    candidate = name
    i = 2
    while candidate.lower() in taken:
        suffix = str(i)
        candidate = name[: 31 - len(suffix)] + suffix
        i += 1
    taken.add(candidate.lower())
    return candidate


def _wallet_columns(db: Session, kind: str) -> list[dict]:
    """All wallets that belong in the workbook, grouped by owner.

    Union of currently active wallets and wallets that appear in snapshots of
    this series (so deleted wallets keep their historical column)."""
    wallets: dict[str, dict] = {}
    for w in db.query(Wallet).filter(Wallet.deleted_at.is_(None)).order_by(Wallet.label).all():
        wallets[w.id] = {"id": w.id, "label": w.label or w.id[:8], "owner": w.owner or "",
                         "wallet_type": w.wallet_type}
    snap_wallets = (
        db.query(BalanceSnapshot.wallet_id, BalanceSnapshot.label, BalanceSnapshot.owner)
        .filter(BalanceSnapshot.kind == kind)
        .distinct()
        .all()
    )
    for wid, label, owner in snap_wallets:
        if wid not in wallets:
            wallets[wid] = {"id": wid, "label": label or wid[:8], "owner": owner or "",
                            "wallet_type": None}
    return sorted(wallets.values(), key=lambda w: (w["owner"] == "", w["owner"].lower(), w["label"].lower()))


def _owner_groups(columns: list[dict]) -> list[dict]:
    groups: list[dict] = []
    for idx, col in enumerate(columns):
        owner = col["owner"]
        if groups and groups[-1]["owner"] == owner:
            groups[-1]["end"] = idx
        else:
            groups.append({"owner": owner, "start": idx, "end": idx})
    return groups


def _overview_rows(db: Session, kind: str) -> list[dict]:
    snaps = (
        db.query(BalanceSnapshot)
        .filter(BalanceSnapshot.kind == kind)
        .order_by(BalanceSnapshot.captured_at, BalanceSnapshot.id)
        .all()
    )
    annotations = {
        (a.kind, a.bucket): a
        for a in db.query(SnapshotAnnotation).filter(SnapshotAnnotation.kind == kind).all()
    }
    buckets: dict[str, dict] = {}
    for s in snaps:
        b = buckets.setdefault(s.bucket, {
            "bucket": s.bucket, "captured_at": s.captured_at, "range_from": s.range_from,
            "range_to": s.range_to, "epoch": s.epoch, "eur_rate": s.eur_rate,
            "usd_rate": s.usd_rate, "trigger": s.trigger, "cells": {},
        })
        b["cells"][s.wallet_id] = s
        if s.eur_rate is not None:
            b["eur_rate"] = s.eur_rate
        if s.usd_rate is not None:
            b["usd_rate"] = s.usd_rate
    rows = sorted(buckets.values(), key=lambda r: (r["captured_at"] or "", r["bucket"]))
    for r in rows:
        ann = annotations.get((kind, r["bucket"]))
        r["why"] = ann.why if ann else None
        r["info"] = ann.info if ann else None
        r["note"] = ann.note if ann else None
    return rows


def _build_overview_sheet(ws, db: Session, kind: str, t: dict):
    columns = _wallet_columns(db, kind)
    groups = _owner_groups(columns)
    rows = _overview_rows(db, kind)

    n = len(columns)
    col_date, col_epoch, col_from, col_to, col_eur, col_usd, col_why = 1, 2, 3, 4, 5, 6, 7
    first_wallet = 8
    col_info = first_wallet + n
    first_owner = col_info + 1
    col_note = first_owner + len(groups)

    r_title, r_total, r_prev, r_cur, r_owner_sum, r_header = 1, 2, 3, 4, 6, 7
    data_start = r_header + 1
    data_end = data_start + max(len(rows), 1) - 1

    bold = Font(bold=True)

    ws.cell(row=r_title, column=1, value=t["title"][kind]).font = bold

    # Column headers
    headers = [
        (col_date, t["date"]), (col_epoch, t["epoch"]), (col_from, t["from"]), (col_to, t["to"]),
        (col_eur, t["eur_per_bln"]), (col_usd, t["usd_per_bln"]), (col_why, t["why"]),
        (col_info, t["info"]), (col_note, t["notes"]),
    ]
    for c, label in headers:
        ws.cell(row=r_header, column=c, value=label).font = bold
    for i, col in enumerate(columns):
        ws.cell(row=r_header, column=first_wallet + i, value=col["label"]).font = bold
    for gi, g in enumerate(groups):
        ws.cell(row=r_header, column=first_owner + gi, value=g["owner"] or t["no_owner"]).font = bold

    # Summary block: previous years (constant), current year (formula), total
    cur_year = datetime.now().year
    ws.cell(row=r_total, column=col_why, value=t["total"]).font = bold
    ws.cell(row=r_prev, column=col_why, value=t["prev_year"]).font = bold
    ws.cell(row=r_cur, column=col_why, value=t["cur_year"]).font = bold

    cur_year_start = data_start
    for idx, row in enumerate(rows):
        dt = _to_local(_parse_iso(row["captured_at"]))
        if dt and dt.year >= cur_year:
            cur_year_start = data_start + idx
            break
    else:
        cur_year_start = data_end + 1  # no current-year rows

    for i, col in enumerate(columns):
        letter = get_column_letter(first_wallet + i)
        prev_sum = 0
        for idx, row in enumerate(rows):
            dt = _to_local(_parse_iso(row["captured_at"]))
            if dt and dt.year < cur_year:
                snap = row["cells"].get(col["id"])
                if snap is not None and snap.delta is not None:
                    prev_sum += snap.delta
        c_prev = ws.cell(row=r_prev, column=first_wallet + i, value=prev_sum)
        c_prev.number_format = FMT_AMOUNT
        if cur_year_start <= data_end:
            c_cur = ws.cell(row=r_cur, column=first_wallet + i,
                            value=f"=SUM({letter}{cur_year_start}:{letter}{data_end})")
        else:
            c_cur = ws.cell(row=r_cur, column=first_wallet + i, value=0)
        c_cur.number_format = FMT_AMOUNT
        c_total = ws.cell(row=r_total, column=first_wallet + i,
                          value=f"={letter}{r_prev}+{letter}{r_cur}")
        c_total.number_format = FMT_AMOUNT

    # Per-owner EUR sum row (above header, like the template)
    for gi in range(len(groups)):
        letter = get_column_letter(first_owner + gi)
        c = ws.cell(row=r_owner_sum, column=first_owner + gi,
                    value=f"=SUM({letter}{data_start}:{letter}{data_end})")
        c.number_format = FMT_EUR
        c.font = bold

    # Data rows
    for idx, row in enumerate(rows):
        r = data_start + idx
        dt = _to_local(_parse_iso(row["captured_at"]))
        if dt:
            c = ws.cell(row=r, column=col_date, value=dt)
            c.number_format = FMT_DATE
        ws.cell(row=r, column=col_epoch, value=row["epoch"])
        d_from = _to_local(_parse_iso(row["range_from"]))
        if d_from:
            c = ws.cell(row=r, column=col_from, value=d_from)
            c.number_format = FMT_DATE
        d_to = _to_local(_parse_iso(row["range_to"]))
        if d_to:
            c = ws.cell(row=r, column=col_to, value=d_to)
            c.number_format = FMT_DATE
        if row["eur_rate"] is not None:
            c = ws.cell(row=r, column=col_eur, value=row["eur_rate"] * BLN)
            c.number_format = FMT_EUR
        if row["usd_rate"] is not None:
            c = ws.cell(row=r, column=col_usd, value=row["usd_rate"] * BLN)
            c.number_format = FMT_USD
        why = row["why"]
        if not why and row["trigger"] == "manual":
            why = t["cat_manual"]
        ws.cell(row=r, column=col_why, value=why)
        ws.cell(row=r, column=col_info, value=row["info"])
        ws.cell(row=r, column=col_note, value=row["note"])

        for i, col in enumerate(columns):
            snap = row["cells"].get(col["id"])
            if snap is not None and snap.delta is not None:
                c = ws.cell(row=r, column=first_wallet + i, value=snap.delta)
                c.number_format = FMT_AMOUNT

        eur_letter = get_column_letter(col_eur)
        for gi, g in enumerate(groups):
            start_letter = get_column_letter(first_wallet + g["start"])
            end_letter = get_column_letter(first_wallet + g["end"])
            c = ws.cell(
                row=r, column=first_owner + gi,
                value=f"=(SUM({start_letter}{r}:{end_letter}{r})/{BLN})*{eur_letter}{r}",
            )
            c.number_format = FMT_EUR

    # Column widths
    ws.column_dimensions[get_column_letter(col_date)].width = 9
    ws.column_dimensions[get_column_letter(col_epoch)].width = 6
    ws.column_dimensions[get_column_letter(col_from)].width = 9
    ws.column_dimensions[get_column_letter(col_to)].width = 9
    ws.column_dimensions[get_column_letter(col_eur)].width = 11
    ws.column_dimensions[get_column_letter(col_usd)].width = 11
    ws.column_dimensions[get_column_letter(col_why)].width = 14
    for i in range(n):
        ws.column_dimensions[get_column_letter(first_wallet + i)].width = 13
    ws.column_dimensions[get_column_letter(col_info)].width = 40
    for gi in range(len(groups)):
        ws.column_dimensions[get_column_letter(first_owner + gi)].width = 11
    ws.column_dimensions[get_column_letter(col_note)].width = 25


def _build_balance_sheet(ws, db: Session, kind: str, t: dict):
    """Absolute balance per wallet and capture (as reported by the network,
    validForTick stored per row) — complements the delta-based Ledger sheet.
    User-entered rows carry no balance and are skipped here."""
    bold = Font(bold=True)
    columns = _wallet_columns(db, kind)
    rows = [
        r for r in _overview_rows(db, kind)
        if any(s.balance is not None for s in r["cells"].values())
    ]

    n = len(columns)
    col_date, col_epoch, col_eur, col_usd = 1, 2, 3, 4
    first_wallet = 5
    col_total = first_wallet + n
    col_val_eur = col_total + 1
    col_val_usd = col_total + 2

    r_title, r_owner, r_header = 1, 2, 3
    data_start = r_header + 1

    ws.cell(row=r_title, column=1, value=t["balance_title"]).font = bold
    for i, col in enumerate(columns):
        ws.cell(row=r_owner, column=first_wallet + i, value=col["owner"] or t["no_owner"]).font = bold
        ws.cell(row=r_header, column=first_wallet + i, value=col["label"]).font = bold
    for c, label in [(col_date, t["date"]), (col_epoch, t["epoch"]),
                     (col_eur, t["eur_per_bln"]), (col_usd, t["usd_per_bln"]),
                     (col_total, t["total"]), (col_val_eur, t["value_eur"]),
                     (col_val_usd, t["value_usd"])]:
        ws.cell(row=r_header, column=c, value=label).font = bold

    for idx, row in enumerate(rows):
        r = data_start + idx
        dt = _to_local(_parse_iso(row["captured_at"]))
        if dt:
            c = ws.cell(row=r, column=col_date, value=dt)
            c.number_format = FMT_DATE
        ws.cell(row=r, column=col_epoch, value=row["epoch"])
        if row["eur_rate"] is not None:
            c = ws.cell(row=r, column=col_eur, value=row["eur_rate"] * BLN)
            c.number_format = FMT_EUR
        if row["usd_rate"] is not None:
            c = ws.cell(row=r, column=col_usd, value=row["usd_rate"] * BLN)
            c.number_format = FMT_USD
        for i, col in enumerate(columns):
            snap = row["cells"].get(col["id"])
            if snap is not None and snap.balance is not None:
                c = ws.cell(row=r, column=first_wallet + i, value=snap.balance)
                c.number_format = FMT_AMOUNT
        start_letter = get_column_letter(first_wallet)
        end_letter = get_column_letter(first_wallet + n - 1)
        c = ws.cell(row=r, column=col_total, value=f"=SUM({start_letter}{r}:{end_letter}{r})")
        c.number_format = FMT_AMOUNT
        total_letter = get_column_letter(col_total)
        eur_letter = get_column_letter(col_eur)
        usd_letter = get_column_letter(col_usd)
        c = ws.cell(row=r, column=col_val_eur, value=f"=({total_letter}{r}/{BLN})*{eur_letter}{r}")
        c.number_format = FMT_EUR
        c = ws.cell(row=r, column=col_val_usd, value=f"=({total_letter}{r}/{BLN})*{usd_letter}{r}")
        c.number_format = FMT_USD

    ws.column_dimensions[get_column_letter(col_date)].width = 9
    ws.column_dimensions[get_column_letter(col_epoch)].width = 6
    ws.column_dimensions[get_column_letter(col_eur)].width = 11
    ws.column_dimensions[get_column_letter(col_usd)].width = 11
    for i in range(n):
        ws.column_dimensions[get_column_letter(first_wallet + i)].width = 14
    ws.column_dimensions[get_column_letter(col_total)].width = 15
    ws.column_dimensions[get_column_letter(col_val_eur)].width = 11
    ws.column_dimensions[get_column_letter(col_val_usd)].width = 11


def _signed_amount(e: Event) -> int:
    amount = e.amount_qubic or 0
    return amount if e.destination_addr == e.wallet_id else -amount


def build_owner_rows(db: Session, wallets: list[dict], kind: str) -> list[dict]:
    """Rows of a per-owner ledger, shared by the Excel sheet and the API.

    Source of truth is the network balance captured per series slot: for each
    capture, the events row is computed as balance delta minus the known TX of
    the interval (locally synced SC events are deliberately NOT used — user
    decision). Regular transfers (TX) stay individual rows dated with the
    transaction's own timestamp. Rows start with the first capture of the
    series (baseline); without captures the ledger is empty.

    Example: previous balance 100, new balance 200, one TX of +50 in between
    → the events row of that slot carries the remaining +50.
    """
    wallet_ids = [w["id"] for w in wallets]

    snaps = (
        db.query(BalanceSnapshot)
        .filter(
            BalanceSnapshot.kind == kind,
            BalanceSnapshot.wallet_id.in_(wallet_ids),
            BalanceSnapshot.trigger.in_(("auto", "manual")),
        )
        .order_by(BalanceSnapshot.captured_at, BalanceSnapshot.id)
        .all()
    )
    if not snaps:
        return []

    start_dt = min(d for d in (_parse_iso(s.captured_at) for s in snaps) if d is not None)

    txs = (
        db.query(Event)
        .filter(Event.wallet_id.in_(wallet_ids), Event.source_type == "TX")
        .order_by(Event.timestamp, Event.tick_number)
        .all()
    )

    rows: list[dict] = []
    # snapshot boundaries per wallet for interval assignment of TX sums
    from bisect import bisect_left
    bounds: dict[str, list[tuple[datetime, str]]] = {}
    for s in snaps:
        dt = _parse_iso(s.captured_at)
        if dt is not None:
            bounds.setdefault(s.wallet_id, []).append((dt, s.bucket))
    for lst in bounds.values():
        lst.sort(key=lambda x: x[0])

    tx_sums: dict[tuple[str, str], int] = {}
    for e in txs:
        dt = _parse_iso(e.timestamp)
        if dt is None or dt <= start_dt:
            continue  # before the series baseline — not part of any interval
        rows.append({
            "type": "tx", "ts": dt, "epoch": e.epoch, "count": 1,
            "cells": {e.wallet_id: _signed_amount(e)},
            "eur_rate": e.qubic_eur_rate, "usd_rate": e.qubic_usd_rate,
            "is_internal": bool(e.is_internal),
            "incoming": e.destination_addr == e.wallet_id,
            "comment": e.comment, "note": e.note, "tx_id": e.id,
        })
        wallet_bounds = bounds.get(e.wallet_id)
        if not wallet_bounds:
            continue
        dts = [b[0] for b in wallet_bounds]
        idx = bisect_left(dts, dt)  # first capture at/after the TX
        if idx < len(wallet_bounds):
            key = (wallet_bounds[idx][1], e.wallet_id)
            tx_sums[key] = tx_sums.get(key, 0) + _signed_amount(e)

    # one events row per capture bucket: delta minus TX of the interval
    buckets: dict[str, dict] = {}
    for s in snaps:
        if s.delta is None:
            continue  # baseline rows have no interval yet
        b = buckets.setdefault(s.bucket, {
            "type": "aggregate", "ts": _parse_iso(s.captured_at), "epoch": s.epoch,
            "count": 0, "cells": {}, "eur_rate": s.eur_rate, "usd_rate": s.usd_rate,
            "is_internal": False, "incoming": True,
            "comment": None, "note": None, "tx_id": None,
        })
        residual = s.delta - tx_sums.get((s.bucket, s.wallet_id), 0)
        if residual:
            b["cells"][s.wallet_id] = residual
            b["count"] += 1
        if b["epoch"] is None and s.epoch is not None:
            b["epoch"] = s.epoch

    rows.extend(b for b in buckets.values() if b["cells"])
    rows.sort(key=lambda r: (r["ts"] or datetime.min, r["type"]))
    return rows


def _row_category(row: dict, t: dict) -> str:
    if row["type"] == "aggregate":
        return t["cat_reward"]
    if row["comment"]:
        return row["comment"]
    if row["is_internal"]:
        return t["cat_internal"]
    return t["cat_in"] if row["incoming"] else t["cat_out"]


def _build_owner_sheet(ws, db: Session, owner: str, wallets: list[dict], t: dict, kind: str):
    bold = Font(bold=True)
    col_date, col_epoch, col_eur, col_usd, col_why = 1, 2, 3, 4, 5
    first_wallet = 6
    col_note = first_wallet + len(wallets)
    r_sum, r_label, r_id = 1, 2, 3
    data_start = r_id + 1

    for c, label in [(col_date, t["date"]), (col_epoch, t["epoch"]), (col_eur, t["eur_per_bln"]),
                     (col_usd, t["usd_per_bln"]), (col_why, t["why"]), (col_note, t["note"])]:
        ws.cell(row=r_label, column=c, value=label).font = bold
    for i, w in enumerate(wallets):
        ws.cell(row=r_label, column=first_wallet + i, value=w["label"]).font = bold
        # full publicId below the label — the UI shows the first 5 chars only
        ws.cell(row=r_id, column=first_wallet + i, value=w["id"])

    rows = build_owner_rows(db, wallets, kind)
    col_of = {w["id"]: first_wallet + i for i, w in enumerate(wallets)}
    r = data_start
    for row in rows:
        if row["ts"]:
            c = ws.cell(row=r, column=col_date, value=_to_local(row["ts"]))
            c.number_format = FMT_DATE
        ws.cell(row=r, column=col_epoch, value=row["epoch"])
        if row["eur_rate"] is not None:
            c = ws.cell(row=r, column=col_eur, value=row["eur_rate"] * BLN)
            c.number_format = FMT_EUR
        if row["usd_rate"] is not None:
            c = ws.cell(row=r, column=col_usd, value=row["usd_rate"] * BLN)
            c.number_format = FMT_USD
        ws.cell(row=r, column=col_why, value=_row_category(row, t))
        for wid, amount in row["cells"].items():
            if wid in col_of:
                c = ws.cell(row=r, column=col_of[wid], value=amount)
                c.number_format = FMT_AMOUNT
        note = row["note"]
        if row["type"] == "aggregate":
            note = t["residual_note"]
        ws.cell(row=r, column=col_note, value=note)
        r += 1

    data_end = max(r - 1, data_start)
    ws.cell(row=r_sum, column=col_why, value=t["total"]).font = bold
    for i in range(len(wallets)):
        letter = get_column_letter(first_wallet + i)
        c = ws.cell(row=r_sum, column=first_wallet + i,
                    value=f"=SUM({letter}{data_start}:{letter}{data_end})")
        c.number_format = FMT_AMOUNT

    ws.column_dimensions[get_column_letter(col_date)].width = 9
    ws.column_dimensions[get_column_letter(col_eur)].width = 11
    ws.column_dimensions[get_column_letter(col_usd)].width = 11
    ws.column_dimensions[get_column_letter(col_why)].width = 16
    for i in range(len(wallets)):
        ws.column_dimensions[get_column_letter(first_wallet + i)].width = 14
    ws.column_dimensions[get_column_letter(col_note)].width = 25


def _build_transfer_sheet(ws, db: Session, columns: list[dict], t: dict, kind: str):
    bold = Font(bold=True)
    col_date = 1
    first_wallet = 2
    r_sum, r_owner, r_label = 1, 2, 3
    data_start = r_label + 1

    ws.cell(row=r_sum, column=col_date, value=t["transfers"]).font = bold
    for i, col in enumerate(columns):
        ws.cell(row=r_owner, column=first_wallet + i, value=col["owner"] or t["no_owner"]).font = bold
        ws.cell(row=r_label, column=first_wallet + i, value=col["label"]).font = bold

    owned = {c["id"] for c in columns}
    col_of = {c["id"]: first_wallet + i for i, c in enumerate(columns)}

    # Only transfers inside the recorded period of this series — everything
    # before the first capture (baseline) predates the ledger, like in
    # build_owner_rows. Without captures the sheet stays empty.
    start_dt = _series_baseline(db, kind)

    internal = (
        db.query(Event)
        .filter(Event.is_internal == 1)
        .order_by(Event.timestamp, Event.tick_number)
        .all()
    )
    seen: set[str] = set()
    r = data_start
    for e in internal:
        if e.id in seen:
            continue  # internal transfers are stored once per involved wallet
        seen.add(e.id)
        if e.source_address not in owned and e.destination_addr not in owned:
            continue
        dt = _parse_iso(e.timestamp)
        if start_dt is None or dt is None or dt <= start_dt:
            continue
        c = ws.cell(row=r, column=col_date, value=_to_local(dt))
        c.number_format = FMT_DATE
        amount = e.amount_qubic or 0
        if e.source_address in col_of:
            c = ws.cell(row=r, column=col_of[e.source_address], value=-amount)
            c.number_format = FMT_AMOUNT
        if e.destination_addr in col_of:
            c = ws.cell(row=r, column=col_of[e.destination_addr], value=amount)
            c.number_format = FMT_AMOUNT
        r += 1

    data_end = max(r - 1, data_start)
    for i in range(len(columns)):
        letter = get_column_letter(first_wallet + i)
        c = ws.cell(row=r_sum, column=first_wallet + i,
                    value=f"=SUM({letter}{data_start}:{letter}{data_end})")
        c.number_format = FMT_AMOUNT

    ws.column_dimensions[get_column_letter(col_date)].width = 9
    for i in range(len(columns)):
        ws.column_dimensions[get_column_letter(first_wallet + i)].width = 13


def build_flat_rows(db: Session, columns: list[dict], kind: str) -> list[dict]:
    """Flat MyLedgerCSV rows (Excel sheet + API): one row per TX with its own
    timestamp, plus one row per capture slot and wallet with the events
    residual (balance delta minus TX) — same source as the owner ledgers."""
    rows = build_owner_rows(db, columns, kind)
    flat: list[dict] = []
    for row in rows:
        if row["type"] == "tx":
            wid, amount = next(iter(row["cells"].items()))
            flat.append({
                "ts": row["ts"], "epoch": row["epoch"], "wallet_id": wid,
                "amount": amount, "eur_rate": row["eur_rate"], "usd_rate": row["usd_rate"],
                "is_sc": False, "is_internal": row["is_internal"],
                "comment": row["comment"], "note": row["note"], "tx_id": row["tx_id"],
            })
        else:
            for wid, amount in row["cells"].items():
                flat.append({
                    "ts": row["ts"], "epoch": row["epoch"], "wallet_id": wid,
                    "amount": amount, "eur_rate": row["eur_rate"], "usd_rate": row["usd_rate"],
                    "is_sc": True, "is_internal": False,
                    "comment": None, "note": None, "tx_id": None,
                })
    flat.sort(key=lambda r: r["ts"] or datetime.min)
    return flat


def _build_transactions_sheet(ws, db: Session, columns: list[dict], t: dict, kind: str):
    bold = Font(bold=True)
    headers = t["csv_headers"]
    for c, label in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=label).font = bold

    owner_of = {c["id"]: c["owner"] for c in columns}
    type_of = {c["id"]: c["wallet_type"] for c in columns}

    r = 2
    for row in build_flat_rows(db, columns, kind):
        if row["ts"]:
            c = ws.cell(row=r, column=1, value=_to_local(row["ts"]))
            c.number_format = 'dd/mm/yyyy'
        ws.cell(row=r, column=2, value="Qubic")
        c = ws.cell(row=r, column=3, value=row["amount"])
        c.number_format = FMT_AMOUNT
        if row["eur_rate"] is not None:
            c = ws.cell(row=r, column=4, value=row["eur_rate"])
            c.number_format = FMT_RATE
        if row["usd_rate"] is not None:
            c = ws.cell(row=r, column=5, value=row["usd_rate"])
            c.number_format = FMT_RATE
        ws.cell(row=r, column=6, value=t["typ_sc"] if row["is_sc"] else t["typ_tx"])
        ws.cell(row=r, column=7, value="")
        ws.cell(row=r, column=8, value="Qubic")
        c = ws.cell(row=r, column=9, value=f"=C{r}*D{r}")
        c.number_format = '#,##0.00'
        wtype = type_of.get(row["wallet_id"])
        ws.cell(row=r, column=10, value=(wtype or "").capitalize() if wtype else "")
        ws.cell(row=r, column=11, value=owner_of.get(row["wallet_id"]) or "")
        text = row["comment"] or (t["text_sc"] if row["is_sc"] else "")
        ws.cell(row=r, column=12, value=text)
        ws.cell(row=r, column=13, value=row["wallet_id"])
        ws.cell(row=r, column=14, value=row["tx_id"] or "")
        r += 1

    widths = [11, 7, 12, 13, 13, 14, 10, 9, 10, 10, 14, 40, 24, 24]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_workbook(db: Session, kind: str, lang: str = "de") -> Workbook:
    t = I18N.get(lang, I18N["de"])
    wb = Workbook()
    taken: set = set()

    ws = wb.active
    ws.title = _sheet_name("Ledger", taken)
    _build_overview_sheet(ws, db, kind, t)

    _build_balance_sheet(wb.create_sheet(_sheet_name(t["balance_sheet"], taken)), db, kind, t)

    columns = _wallet_columns(db, kind)
    groups = _owner_groups(columns)
    for g in groups:
        owner = g["owner"]
        wallets = columns[g["start"]:g["end"] + 1]
        name = _sheet_name(f"Ledger{owner or t['no_owner']}", taken)
        _build_owner_sheet(wb.create_sheet(name), db, owner, wallets, t, kind)

    _build_transfer_sheet(wb.create_sheet(_sheet_name("Transfer", taken)), db, columns, t, kind)
    _build_transactions_sheet(wb.create_sheet(_sheet_name("MyLedgerCSV", taken)), db, columns, t, kind)
    return wb


def build_workbook_bytes(db: Session, kind: str, lang: str = "de") -> bytes:
    wb = build_workbook(db, kind, lang)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_dir() -> Path:
    url = settings.database_url
    if "sqlite" in url:
        db_path = Path(url.split("sqlite:///")[-1])
        base = db_path.parent
    else:
        base = Path(".")
    d = base / "exports"
    d.mkdir(parents=True, exist_ok=True)
    return d


def export_file_path(kind: str, lang: str = "de") -> Path:
    names = FILE_NAMES.get(lang, FILE_NAMES["de"])
    return export_dir() / f"qubicflow_bestandsverlauf_{names[kind]}.xlsx"


def write_workbooks(kinds: list[str] | None = None) -> list[str]:
    """Regenerate the export files for all enabled series. Returns written paths.
    A file locked by Excel just logs a warning — the next run rewrites it."""
    from .balance_snapshot_service import KINDS, get_bh_settings

    db = SessionLocal()
    written: list[str] = []
    try:
        settings_map = get_bh_settings(db)
        lang = settings_map.get("bh_export_lang", "de")
        targets = kinds or [k for k in KINDS if settings_map.get(f"bh_{k}_enabled", "true") == "true"]
        for kind in targets:
            path = export_file_path(kind, lang)
            tmp = path.with_suffix(f".tmp{os.getpid()}")
            try:
                # Write to a temp file first, then swap atomically — readers
                # (or Excel) never see a half-written workbook.
                build_workbook(db, kind, lang).save(tmp)
                tmp.replace(path)
                written.append(str(path))
            except PermissionError:
                logger.warning(f"excel export {kind}: file locked (open in Excel?) — skipped this run")
            except Exception as e:
                logger.error(f"excel export {kind} failed: {e}")
            finally:
                tmp.unlink(missing_ok=True)
    finally:
        db.close()
    if written:
        logger.info(f"excel export written: {', '.join(written)}")
    return written
